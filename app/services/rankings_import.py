"""Parses uploaded rankings files (CSV / Excel / PDF) into ranking sets.

Best-effort extraction: we look for a rank number and a player name per row
(plus optional position/team columns), then match names against the real
Sleeper player pool. Unmatched rows are reported back to the user rather
than silently dropped.
"""
from __future__ import annotations

import csv
import io
import re

from app.services import rankings_store

POSITION_TOKENS = {"QB", "RB", "WR", "TE", "K", "DST", "DEF", "D/ST"}


def _parse_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(c.strip() for c in row)]


def _parse_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if v is None else str(v).strip() for v in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


_PDF_LINE = re.compile(
    r"^\s*(\d{1,3})[.)\s]+([A-Za-z][A-Za-z.'\- ]+?)"
    r"(?:[,\s]+(QB|RB|WR|TE|K|DST|DEF|D/ST))?"
    r"(?:[,\s]+([A-Z]{2,3}))?\s*$"
)


def _parse_pdf(content: bytes) -> list[dict]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    entries = []
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            m = _PDF_LINE.match(line.strip())
            if not m:
                continue
            rank, name, pos, team = m.groups()
            entries.append({"rank": int(rank), "name": name.strip(), "pos": pos, "team": team})
    return entries


def _find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]] | None:
    """Returns (header_row_index, {field: column_index}) if a header is found."""
    for i, row in enumerate(rows[:5]):
        lowered = [c.lower() for c in row]
        cols: dict[str, int] = {}
        for j, cell in enumerate(lowered):
            if "name" not in cols and ("player" in cell or cell == "name"):
                cols["name"] = j
            if "rank" not in cols and ("rank" in cell or cell in ("rk", "#", "ovr", "overall")):
                cols["rank"] = j
            if "pos" not in cols and cell.startswith("pos"):
                cols["pos"] = j
            if "team" not in cols and cell in ("team", "tm", "nfl team"):
                cols["team"] = j
            if "sos" not in cols and "sos" in cell:
                cols["sos"] = j
            if "bye" not in cols and "bye" in cell:
                cols["bye"] = j
        if "name" in cols:
            return i, cols
    return None


def _rows_to_entries(rows: list[list[str]]) -> list[dict]:
    header = _find_header(rows)
    entries = []
    if header:
        header_idx, cols = header
        for row in rows[header_idx + 1:]:
            if cols["name"] >= len(row):
                continue
            name = row[cols["name"]].strip()
            if not name:
                continue
            rank = None
            if "rank" in cols and cols["rank"] < len(row):
                m = re.search(r"\d+", row[cols["rank"]])
                if m:
                    rank = int(m.group())
            pos = row[cols["pos"]].strip().upper() if "pos" in cols and cols["pos"] < len(row) else None
            team = row[cols["team"]].strip().upper() if "team" in cols and cols["team"] < len(row) else None
            sos = None
            if "sos" in cols and cols["sos"] < len(row):
                m2 = re.search(r"([0-5])", str(row[cols["sos"]]))
                if m2:
                    sos = int(m2.group(1))
            bye = None
            if "bye" in cols and cols["bye"] < len(row):
                m3 = re.search(r"\d{1,2}", str(row[cols["bye"]]))
                if m3:
                    bye = int(m3.group())
            entries.append({"rank": rank, "name": name, "pos": pos or None, "team": team or None,
                            "sos": sos, "bye": bye})
    else:
        # No header: assume "rank, name, ..." or "name, ..." per row.
        for row in rows:
            cells = [c.strip() for c in row if c.strip()]
            if not cells:
                continue
            if re.fullmatch(r"\d{1,3}", cells[0]) and len(cells) > 1:
                entries.append({"rank": int(cells[0]), "name": cells[1], "pos": None, "team": None})
            else:
                entries.append({"rank": None, "name": cells[0], "pos": None, "team": None})

    for i, e in enumerate(entries):
        if e["rank"] is None:
            e["rank"] = i + 1
    return entries


async def parse_and_match(filename: str, content: bytes) -> dict:
    lowered = filename.lower()
    if lowered.endswith(".csv") or lowered.endswith(".tsv") or lowered.endswith(".txt"):
        entries = _rows_to_entries(_parse_csv(content))
    elif lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        entries = _rows_to_entries(_parse_xlsx(content))
    elif lowered.endswith(".pdf"):
        entries = _parse_pdf(content)
    else:
        raise ValueError("Unsupported file type — use .csv, .xlsx, or .pdf")

    if not entries:
        raise ValueError("No ranking rows could be extracted from this file")

    name_index = await rankings_store.build_name_index()

    matched: dict[str, dict] = {}
    unmatched: list[dict] = []
    for e in entries:
        raw_name = e["name"]
        # Normalize FP-style position values ("RB1", "DST3") to plain codes.
        if e.get("pos"):
            pos = re.sub(r"\d+$", "", e["pos"].strip().upper())
            e["pos"] = {"DST": "DEF", "D/ST": "DEF", "PK": "K"}.get(pos, pos) or None
        # Strip a trailing position/team glued onto the name ("Bijan Robinson RB ATL")
        cleaned = re.sub(
            r"\s+(QB|RB|WR|TE|K|DST|DEF|D/ST)(\s+[A-Z]{2,3})?$", "", raw_name
        ).strip()

        pid = None
        def_abbr = rankings_store.match_defense(raw_name, e.get("pos"))
        if def_abbr:
            pid = def_abbr
        else:
            candidates = name_index.get(rankings_store.normalize_name(cleaned), [])
            if e.get("pos") and len(candidates) > 1:
                pos = e["pos"].replace("DST", "DEF").replace("D/ST", "DEF")
                filtered = [c for c in candidates if c["position"] == pos]
                candidates = filtered or candidates
            if candidates:
                # On collisions prefer the player the market cares about most.
                candidates.sort(key=lambda c: c.get("market_rank") or 10**9)
                pid = candidates[0]["id"]

        if pid is None:
            unmatched.append({"rank": e["rank"], "name": raw_name})
        elif pid not in matched or e["rank"] < matched[pid]["rank"]:
            matched[pid] = {"player_id": pid, "rank": e["rank"], "name": cleaned,
                            "sos": e.get("sos"), "bye": e.get("bye")}

    ranked = sorted(matched.values(), key=lambda m: m["rank"])
    return {"matched": ranked, "unmatched": unmatched}
